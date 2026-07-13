# -*- coding: utf-8 -*-
"""Excel志愿表生成器 - analysis.json → .xlsx（openpyxl）

生成3个Sheet：
1. 志愿清单 — 96个志愿完整表格（序号/院校/专业/冲稳保/位次/学费）
2. 学生画像 — 基本信息+学科成绩+推荐方向
3. 风险评估 — 滑档/退档风险+确认清单
"""

import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# -- 样式常量 --
DEEP_BLUE = "1E3A5F"
BURGUNDY = "6E2C2C"
DARK_GREEN = "2D4A3E"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
DARK_TEXT = "1a1a1a"

thin_border = Border(
    left=Side(style="thin", color="AAAAAA"),
    right=Side(style="thin", color="AAAAAA"),
    top=Side(style="thin", color="AAAAAA"),
    bottom=Side(style="thin", color="AAAAAA"),
)

header_font = Font(name="等线", size=11, bold=True, color=WHITE)
header_fill = PatternFill(start_color=DEEP_BLUE, end_color=DEEP_BLUE, fill_type="solid")
body_font = Font(name="等线", size=10.5, color=DARK_TEXT)
tier_font = Font(name="等线", size=11, bold=True, color=WHITE)
tier_fills = {
    "冲": PatternFill(start_color=BURGUNDY, end_color=BURGUNDY, fill_type="solid"),
    "稳": PatternFill(start_color=DEEP_BLUE, end_color=DEEP_BLUE, fill_type="solid"),
    "保": PatternFill(start_color=DARK_GREEN, end_color=DARK_GREEN, fill_type="solid"),
}
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)


def load_analysis(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border


def _style_body_cell(cell, alignment=None):
    cell.font = body_font
    cell.alignment = alignment or center
    cell.border = thin_border


def build_sheet_volunteers(wb, analysis):
    """Sheet 1: 志愿清单"""
    ws = wb.create_sheet("志愿清单", 0)
    volunteers = analysis.get("volunteers") or []
    summary = analysis.get("student_summary") or {}

    # 标题行
    ws.merge_cells("A1:G1")
    title = ws.cell(row=1, column=1, value="高考志愿填报参考列表")
    title.font = Font(name="等线", size=16, bold=True, color=DEEP_BLUE)
    title.alignment = center
    ws.row_dimensions[1].height = 32

    # 信息行
    ws.merge_cells("A2:G2")
    info = ws.cell(row=2, column=1,
                   value=f"考生：{summary.get('name', '')}　|　"
                         f"位次：{summary.get('rank', '')}　|　"
                         f"王牌：{summary.get('king_card', '无')}　|　"
                         f"共{len(volunteers)}个志愿")
    info.font = Font(name="等线", size=10, color="888888")
    info.alignment = center
    ws.row_dimensions[2].height = 22

    # 表头
    headers = ["序号", "院校", "专业", "冲稳保", "2025位次", "学费(元/年)", "备注"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    _style_header_row(ws, 4, len(headers))
    ws.row_dimensions[4].height = 28

    # 数据行
    tier_names = {"冲": "冲档", "稳": "稳档", "保": "保档"}
    current_tier = None
    row = 5

    for v in volunteers:
        tier = v.get("tier", "")
        # 分档标题行
        if tier != current_tier:
            current_tier = tier
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            cell = ws.cell(row=row, column=1, value=f"━━ {tier_names.get(tier, tier)} ━━")
            cell.font = tier_font
            cell.fill = tier_fills.get(tier, header_fill)
            cell.alignment = center
            cell.border = thin_border
            ws.row_dimensions[row].height = 24
            row += 1

        values = [
            v.get("order", ""),
            v.get("name", ""),
            v.get("major", ""),
            tier,
            v.get("rank_2025", ""),
            v.get("tuition", ""),
            v.get("note", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            _style_body_cell(cell, left if col in (2, 3, 7) else center)
        ws.row_dimensions[row].height = 22
        row += 1

    # 底部提示
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    tip = ws.cell(row=row, column=1,
                  value="提示：所有志愿建议勾选「专业服从调剂」。调剂只在同专业组内进行，安全。")
    tip.font = Font(name="等线", size=9, color="888888")
    tip.alignment = left

    # 列宽
    widths = [6, 22, 26, 8, 12, 14, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_sheet_profile(wb, analysis):
    """Sheet 2: 学生画像"""
    ws = wb.create_sheet("学生画像")
    summary = analysis.get("student_summary") or {}
    strengths = analysis.get("strengths") or {}
    directions = analysis.get("directions") or []
    scores = strengths.get("all_scores") or {}

    # 标题
    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value="考生画像与学科优势分析")
    t.font = Font(name="等线", size=16, bold=True, color=DEEP_BLUE)
    t.alignment = center
    ws.row_dimensions[1].height = 32

    # 基本信息
    row = 3
    ws.cell(row=row, column=1, value="基本信息")
    ws.cell(row=row, column=1).font = Font(name="等线", size=12, bold=True, color=WHITE)
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).alignment = center
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1

    basic_items = [
        ("姓名", summary.get("name", "")),
        ("位次", summary.get("rank", "")),
        ("王牌科目", summary.get("king_card") or "无显著王牌"),
        ("软肋科目", strengths.get("weakest") or "无明显软肋"),
        ("推荐方向", "、".join(summary.get("best_directions", []))),
    ]
    for label, val in basic_items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = Font(name="等线", size=10.5, bold=True, color=DARK_TEXT)
        ws.cell(row=row, column=1).fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = center
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.cell(row=row, column=2, value=str(val))
        _style_body_cell(ws.cell(row=row, column=2), left)
        ws.cell(row=row, column=3).border = thin_border
        row += 1

    # 学科成绩
    row += 1
    ws.cell(row=row, column=1, value="学科成绩")
    ws.cell(row=row, column=1).font = Font(name="等线", size=12, bold=True, color=WHITE)
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).alignment = center
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1

    ws.cell(row=row, column=1, value="科目")
    ws.cell(row=row, column=2, value="分数")
    ws.cell(row=row, column=3, value="评估")
    _style_header_row(ws, row, 3)
    row += 1

    for subj, score in sorted(scores.items(), key=lambda x: x[1], reverse=True):
        if score >= 85:
            eval_text = "★ 王牌"
        elif score <= 70:
            eval_text = "△ 软肋"
        else:
            eval_text = "— 正常"
        for col, val in enumerate([subj, score, eval_text], 1):
            cell = ws.cell(row=row, column=col, value=val)
            _style_body_cell(cell)
        row += 1

    # 推荐方向
    row += 1
    ws.cell(row=row, column=1, value="推荐方向与匹配度")
    ws.cell(row=row, column=1).font = Font(name="等线", size=12, bold=True, color=WHITE)
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).alignment = center
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1

    ws.cell(row=row, column=1, value="方向")
    ws.cell(row=row, column=2, value="匹配度")
    ws.cell(row=row, column=3, value="推荐理由")
    _style_header_row(ws, row, 3)
    row += 1

    for d in directions:
        ws.cell(row=row, column=1, value=d.get("direction", ""))
        ws.cell(row=row, column=2, value=f'{d.get("match_score", 0)}%')
        ws.cell(row=row, column=3, value=d.get("reason", ""))
        for col in range(1, 4):
            _style_body_cell(ws.cell(row=row, column=col), left if col == 3 else center)
        row += 1

    # 列宽
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 40


def build_sheet_risk(wb, analysis):
    """Sheet 3: 风险评估"""
    ws = wb.create_sheet("风险评估")
    risks = analysis.get("risk_assessment") or {}
    summary = analysis.get("student_summary") or {}

    # 标题
    ws.merge_cells("A1:B1")
    t = ws.cell(row=1, column=1, value="风险评估与防范")
    t.font = Font(name="等线", size=16, bold=True, color=DEEP_BLUE)
    t.alignment = center
    ws.row_dimensions[1].height = 32

    row = 3
    # 风险等级
    risk_items = [
        ("滑档风险", risks.get("slide_risk", "未评估")),
        ("退档风险", risks.get("reject_risk", "未评估")),
        ("保底位次", risks.get("bottom_rank", "未知")),
    ]
    for label, val in risk_items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = Font(name="等线", size=10.5, bold=True, color=DARK_TEXT)
        ws.cell(row=row, column=1).fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=2, value=str(val))
        _style_body_cell(ws.cell(row=row, column=2), left)
        row += 1

    # 退档原因
    row += 1
    reasons = risks.get("reject_reasons", [])
    if reasons:
        ws.cell(row=row, column=1, value="退档风险因素")
        ws.cell(row=row, column=1).font = Font(name="等线", size=12, bold=True, color=WHITE)
        ws.cell(row=row, column=1).fill = PatternFill(start_color=BURGUNDY, end_color=BURGUNDY, fill_type="solid")
        ws.cell(row=row, column=1).alignment = center
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1
        for r in reasons:
            ws.cell(row=row, column=1, value=f"⚠ {r}")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            _style_body_cell(ws.cell(row=row, column=1), left)
            row += 1

    # 确认清单
    row += 1
    ws.cell(row=row, column=1, value="填报前紧急确认清单")
    ws.cell(row=row, column=1).font = Font(name="等线", size=12, bold=True, color=WHITE)
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).alignment = center
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    ws.cell(row=row, column=1, value="序号")
    ws.cell(row=row, column=2, value="确认事项")
    _style_header_row(ws, row, 2)
    row += 1

    for i, item in enumerate(risks.get("checklist", []), 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=item)
        _style_body_cell(ws.cell(row=row, column=1), center)
        _style_body_cell(ws.cell(row=row, column=2), left)
        row += 1

    # 列宽
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 50


def generate_excel(analysis_path: str, output_xlsx: str):
    """从analysis.json生成Excel志愿表"""
    try:
        analysis = load_analysis(analysis_path)
    except FileNotFoundError:
        print(f"ERROR: 分析结果文件不存在: {analysis_path}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"ERROR: analysis.json格式错误: {e}")
        sys.exit(1)

    output_dir = os.path.dirname(output_xlsx)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    try:
        wb = Workbook()
        wb.remove(wb.active)

        build_sheet_volunteers(wb, analysis)
        build_sheet_profile(wb, analysis)
        build_sheet_risk(wb, analysis)

        wb.properties.title = "高考志愿填报参考列表"
        wb.properties.creator = "gaokao-advisor"
        wb.properties.created = datetime.now()

        wb.save(output_xlsx)
        size = os.path.getsize(output_xlsx)
        print(f"Excel generated: {output_xlsx} ({size:,} bytes)")
    except Exception as e:
        print(f"ERROR: Excel生成失败: {e}")
        sys.exit(1)


if __name__ == "__main__":
    import sys
    if len(sys.argv) >= 3:
        generate_excel(sys.argv[1], sys.argv[2])
    else:
        print("Usage: python excel_generator.py <analysis.json> <output.xlsx>")
